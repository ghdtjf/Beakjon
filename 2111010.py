import requests
from pprint import pprint
from bs4 import BeautifulSoup

# class="rank_number
# class="artist"
# 

cookies = {
    'PC_PCID': '16365289427260957457690',
    'PCID': '16365289427260957457690',
    '__T_': '1',
    'POC': 'WP10',
    '_T_ANO': 'koCnigCly/3b1vmn1jQDt7TdbQ/DZGM2e6TSTaRGMOIAKenTKFkoUPXCQIbilR5wbGmFdNuSzv0sh9F4maQpJiO9/rbwv+08jdivCZA99G+qlDRPkO1m9Wo6YpCQYq9pdXpJiIrmq/ZJTkkpmOtm85izWHLOGsH9Iu7aK4evfFiQuBKfIsUmF/lmkTvH71dlfsuBa8JgKj4gaM3ZMT8sPOqFLeFTN+JGKLRe8T7a3OTi0Cnon+RREMc2UTiSV4YafzEnPCVUZTExb3TNwl7JN824KG8H7q7z+2OMkPhgtH2oLRnR/apTg/wrDT5VBoQToGgjrwAf50jDfnan2NHEmw==',
}

headers = {
    'Connection': 'keep-alive',
    'Cache-Control': 'max-age=0',
    'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'Sec-Fetch-Site': 'cross-site',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-User': '?1',
    'Sec-Fetch-Dest': 'document',
    'Referer': 'https://www.google.com/',
    'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
    'If-None-Match': '"0:6290"',
}

res = requests.get('https://www.melon.com/', headers=headers, cookies=cookies)

bsOdject = BeautifulSoup(res.content, 'lxml')
bsOdject = bsOdject.find("div", { "class": "list_wrap typeRealtime" }).find("ul")

result_list = list()

for item in bsOdject.find_all("li", {"class" : "rank_item"}):
    rank = item.find("span", {"class" : "rank"})
    song = item.find("a", {"class" : "ellipsis mlog"})
    artist = item.find("a", {"class" : "fc_mgray mlog"})
    result_list.append([rank.text, song.text, artist.text])

from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

row_idx = 1

for item in result_list:
    rank_cell = sheet.cell(row=row_idx, column=1)
    song_cell = sheet.cell(row=row_idx, column=2)
    artist_cell = sheet.cell(row=row_idx, column=3)
    rank_cell.value = item[0]
    song_cell.value = item[1]
    artist_cell.value = item[2]
    row_idx += 1

wb.save("멜론차트1.xlsx")